import openpyxl
import os


class WriteExcel(object):  # 写入excel数据的类
    data = []
    file_name = ''

    def __init__(self, file_name, file_data):
        self.data = file_data
        self.file_name = file_name
        """
        这个是用来初始化读取对象的
        :param file_name: 文件名 ---> str类型
        """
        # 打开文件
        file_exist = os.path.exists(file_name)
        if file_exist is False:
            self.wb = openpyxl.Workbook()
            self.wb.save(file_name)
        self.wb = openpyxl.load_workbook(file_name)
        # self.wb = openpyxl.Workbook()
        # 选择表单
        self.sh = self.wb.active

    def write_data_line(self):
        # col = self.sh.max_column
        # row = self.sh.max_row
        for i in self.data:
            self.sh.append(i)
        # if row > 1:
        #     row = row+1
        # else:
        #     row = row
        # for r in range(len(self.data)):
        #     for c in range(len(self.data[r])):
        #         self.sh.cell(row=r + row, column=c + 1, value=self.data[r][c])
        self.wb.save(self.file_name)
        self.wb.close()


#
# if __name__ == '__main__':
#     data = [['as', 'b', 'c', 'd', 'e', 'f', 'g'], [1, 2, 3, 4, 5, 6, 1], [4, 5, 6, 7, 8, 9, 0], [4, 67, 3, 2, 6, 8, 9],[2,4,6,7,8,9]]
#     rw = WriteExcel('excel01.xlsx',data)
#     rw.write_data_line()
