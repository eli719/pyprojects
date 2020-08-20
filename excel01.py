import openpyxl

# 创建一个工作簿,默认打开的文件为可读写，若有需要可以指定参数read_only为True。
wb = openpyxl.Workbook()
# 根据sheet名字获得sheet
ws = wb.get_sheet_by_name('Sheet1')
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
ws = wb.active
# 默认已有一个名为sheet的表单,修改title即修改了表单名称
data = [['a', 'b', 'c', 'd', 'e', 'f', 'g'], [1, 2, 3, 4, 5, 6, 1], [4,5,6,7,8,9,0], [4,67,3,2,6,8,9]]
ws.title = 'sheet01'
print(ws.max_row)
print(ws.max_column)
# 写入excel
print(list(data[1:]))
for r in range(len(data)):
    for c in range(len(data[r])):
        ws.cell(row=r+1, column=c+1, value=data[r][c])
print(ws.max_row)
print(ws.max_column)
# 创建一个新表单，可以指定位置
sheet = wb.create_sheet('new_sheet',1)
# 保存为一个xlsx格式的文件
wb.save('excel01.xlsx')
# 读取excel中的数据
# 打开工作簿
wb = openpyxl.load_workbook('excel01.xlsx')
# 选取表单
ws = wb['sheet01']
# 读取数据
cell = ws.cell(row=1,column=1)  # 读取第一行，第一列数据
print(cell.value)
# 按行读取数据
print(list(ws.rows)[1:])  # 按行读取数据，去掉第一行的表头信息
for cases in list(ws.rows)[1:]:
    for i in range(len(cases)):
        print(i,cases[i].value)
wb.close()